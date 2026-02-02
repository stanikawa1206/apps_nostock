# -*- coding: utf-8 -*-
r"""
Amazon Orders + Finances → Excel 出力（内訳完全版）
出力先: Y:\Amazon輸出\注文data.xlsx
"""

import os
import sys
import requests
import json
import time
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from collections import defaultdict
from dotenv import load_dotenv
from requests_auth_aws_sigv4 import AWSSigV4
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from pathlib import Path
from dateutil import parser

# =========================
# 環境変数
# =========================
# カレントディレクトリの .env を読み込む
load_dotenv()
# スクリプト配置ディレクトリの .env も念のため読み込む (優先度高)
script_env_path = Path(__file__).parent / ".env"
if script_env_path.exists():
    load_dotenv(dotenv_path=script_env_path)

AWS_ACCESS_KEY = os.getenv("AWS_ACCESS_KEY")
AWS_SECRET_KEY = os.getenv("AWS_SECRET_KEY")

LWA_CLIENT_ID = os.getenv("LWA_CLIENT_ID")
LWA_CLIENT_SECRET = os.getenv("LWA_CLIENT_SECRET")

# US設定のみを使用（保険的なフォールバックは廃止）
REFRESH_TOKEN = os.getenv("REFRESH_TOKEN_US")
MARKETPLACE_ID = os.getenv("MARKETPLACE_ID_US")
API_ENDPOINT = os.getenv("SPAPI_ENDPOINT_US", "https://sellingpartnerapi-na.amazon.com")
AWS_REGION = os.getenv("AWS_REGION_US", "us-east-1")

EXCEL_PATH = r"Y:\Amazon輸出\注文data.xlsx"
SHEET_NAME = "amazon"

# Excelのヘッダー定義 (buyer-nameを削除、finance-status, last-updatedを追加)
HEADERS = [
    "amazon-order-id", "purchase-date", "order-status", "asin", "quantity",
    "principal", "shipping-price",
    "gross-sales", "amazon-fee", "net-payout",
    "ship-city", "ship-state", "ship-postal-code", "ship-country", "recipient-name",
    "currency", "order-item-id", "fulfillment", "finance-status", "last-updated"
]

# 必須変数のチェック
if not REFRESH_TOKEN:
    print(f"[ERROR] 環境変数 'REFRESH_TOKEN_US' が見つかりません。")
    print(f"設定ファイル (.env) を確認してください。\n検索パス: {os.getcwd()} または {script_env_path}")
    sys.exit(1)

# =========================
# 共通：LWA Token
# =========================
def get_lwa_token():
    r = requests.post(
        "https://api.amazon.com/auth/o2/token",
        data={
            "grant_type": "refresh_token",
            "refresh_token": REFRESH_TOKEN,
            "client_id": LWA_CLIENT_ID,
            "client_secret": LWA_CLIENT_SECRET,
        },
    )
    if r.status_code != 200:
        print(f"[FATAL] LWA Token取得失敗: {r.status_code}")
        print(f"Response: {r.text}")
        print(f"DEBUG: REFRESH_TOKEN={'Set' if REFRESH_TOKEN else 'None'}")
        print(f"DEBUG: LWA_CLIENT_ID={'Set' if LWA_CLIENT_ID else 'None'}")
        print(f"DEBUG: LWA_CLIENT_SECRET={'Set' if LWA_CLIENT_SECRET else 'None'}")

    r.raise_for_status()
    return r.json()["access_token"]

def get_auth():
    return AWSSigV4(
        "execute-api",
        region=AWS_REGION,
        aws_access_key_id=AWS_ACCESS_KEY,
        aws_secret_access_key=AWS_SECRET_KEY,
    )

def get_open_event_group_id(token, auth):
    r = requests.get(f"{API_ENDPOINT}/finances/v0/financialEventGroups", 
                     headers={"x-amz-access-token": token}, auth=auth)
    if r.status_code == 200:
        groups = r.json().get("payload", {}).get("FinancialEventGroupList", [])
        for g in groups:
            if g.get("ProcessingStatus") == "Open":
                return g.get("FinancialEventGroupId")
    return None

def get_rdt_token(lwa_token, auth):
    """個人情報(PII)アクセス用のRDTトークンを取得"""
    body = {
        "restrictedResources": [
            {
                "method": "GET",
                "path": "/orders/v0/orders",
                "dataElements": ["buyerInfo", "shippingAddress"]
            }
        ]
    }
    r = requests.post(
        f"{API_ENDPOINT}/tokens/2021-03-01/restrictedDataToken",
        headers={"x-amz-access-token": lwa_token},
        auth=auth,
        json=body
    )
    if r.status_code == 200:
        return r.json()["restrictedDataToken"]
    else:
        print(f"[WARN] RDT取得失敗: {r.status_code} {r.text} (個人情報は取得できません)")
        return lwa_token  # フォールバック

# =========================
# Orders API
# =========================
def get_fees_estimate(asin, price, currency, is_fba, token, auth):
    """Product Fees APIを使って概算手数料を取得"""
    if not asin or not price or price <= 0:
        return Decimal("0")

    body = {
        "FeesEstimateRequest": {
            "MarketplaceId": MARKETPLACE_ID,
            "IsAmazonFulfilled": is_fba,
            "PriceToEstimateFees": {
                "ListingPrice": {
                    "CurrencyCode": currency,
                    "Amount": float(price)
                }
            },
            "Identifier": str(asin)
        }
    }

    try:
        r = requests.post(
            f"{API_ENDPOINT}/products/fees/v0/items/{asin}/feesEstimate",
            headers={"x-amz-access-token": token},
            auth=auth,
            json=body
        )
        if r.status_code == 200:
            payload = r.json().get("payload", {})
            result = payload.get("FeesEstimateResult", {})
            if result.get("Status") == "Success":
                detail = result.get("FeesEstimate", {}).get("FeeDetailList", [])
                total = sum(Decimal(str(d.get("FeeAmount", {}).get("Amount", 0))) for d in detail)
                return total * Decimal("-1") # 手数料はマイナスとして扱う
    except Exception:
        pass
    
    return Decimal("0")

def fetch_orders(created_after):
    lwa_token = get_lwa_token()
    auth = get_auth()
    
    # RDTトークンに切り替え（PII取得のため）
    token = get_rdt_token(lwa_token, auth)

    orders = []
    next_token = None

    while True:
        params = {"MarketplaceIds": MARKETPLACE_ID}
        if next_token:
            params["NextToken"] = next_token
        else:
            params["CreatedAfter"] = created_after

        r = requests.get(
            f"{API_ENDPOINT}/orders/v0/orders",
            headers={"x-amz-access-token": token},
            params=params,
            auth=auth,
        )
        
        if r.status_code != 200:
            print(f"[ERROR] Orders API Error: {r.status_code}")
            print(f"URL: {r.url}")
            print(f"Response: {r.text}")
            print("※ 403エラーの場合、セラーセントラルでアプリの権限(ロール)を確認するか、マーケットプレイスID(JP/US)が正しいか確認してください。")
            r.raise_for_status()

        payload = r.json()["payload"]
        orders.extend(payload.get("Orders", []))
        next_token = payload.get("NextToken")
        if not next_token:
            break

    return orders

def fetch_order_items(order_id, token, auth):
    r = requests.get(
        f"{API_ENDPOINT}/orders/v0/orders/{order_id}/orderItems",
        headers={"x-amz-access-token": token},
        auth=auth,
    )
    r.raise_for_status()
    return r.json()["payload"]["OrderItems"]

# =========================
# Finances API（内訳取得）
# =========================
def fetch_finances(days=180):
    token = get_lwa_token()
    auth = get_auth()

    # 1. Open Group ID (未確定分)
    open_group_id = get_open_event_group_id(token, auth)

    # 2. 期間指定 (確定分)
    posted_after = (
        datetime.now(timezone.utc) - timedelta(days=days)
    ).strftime("%Y-%m-%dT%H:%M:%SZ")

    targets = []
    if open_group_id:
        targets.append({"FinancialEventGroupId": open_group_id})
    targets.append({"PostedAfter": posted_after})

    rows = []
    service_fees = []

    for params_base in targets:
        next_token = None
        while True:
            params = params_base.copy()
            if next_token:
                params = {"NextToken": next_token}

            r = requests.get(
                f"{API_ENDPOINT}/finances/v0/financialEvents",
                headers={"x-amz-access-token": token},
                params=params,
                auth=auth,
            )
            if r.status_code != 200:
                print(f"[WARN] fetch_finances failed: {r.status_code} {r.text}")
                break

            payload = r.json().get("payload", {})
            events = payload.get("FinancialEvents", {})

            # Shipment
            for ev in events.get("ShipmentEventList", []):
                for item in ev.get("ShipmentItemList", []):
                    rows.append(parse_finance_item(
                        item.get("OrderItemId"),
                        item.get("ItemChargeList", []),
                        item.get("ItemFeeList", []),
                        item.get("ItemTaxWithheldList", [])
                    ))

            # Refund
            for ev in events.get("RefundEventList", []):
                for item in ev.get("ShipmentItemAdjustmentList", []):
                    rows.append(parse_finance_item(
                        item.get("OrderItemId"),
                        item.get("ItemChargeAdjustmentList", []),
                        item.get("ItemFeeAdjustmentList", []),
                        item.get("ItemTaxWithheldList", [])
                    ))

            # Service Fees (月額登録料、広告費など)
            for ev in events.get("ServiceFeeEventList", []):
                fee_obj = ev.get("FeeAmount") or {}
                service_fees.append({
                    "date": ev.get("PostedDate"),
                    "type": "ServiceFee",
                    "description": ev.get("FeeDescription"),
                    "amount": ev.get("FeeAmount", {}).get("CurrencyAmount", 0),
                    "amount": fee_obj.get("CurrencyAmount", 0),
                })

            next_token = payload.get("NextToken")
            if not next_token:
                break

    return rows, service_fees

def parse_finance_item(order_item_id, charge_list, fee_list, tax_withheld_list=None):
    if tax_withheld_list is None:
        tax_withheld_list = []

    charges = defaultdict(Decimal)
    for c in charge_list:
        amt = c.get("ChargeAmount", {}).get("CurrencyAmount", 0)
        c_amt = c.get("ChargeAmount") or {}
        amt = c_amt.get("CurrencyAmount", 0)
        charges[c["ChargeType"]] += Decimal(str(amt))

    fees = defaultdict(Decimal)
    for f in fee_list:
        amt = f.get("FeeAmount", {}).get("CurrencyAmount", 0)
        f_amt = f.get("FeeAmount") or {}
        amt = f_amt.get("CurrencyAmount", 0)
        fees[f["FeeType"]] += Decimal(str(amt))

    # Tax Withheld (Marketplace Facilitator Taxなど)
    tax_withheld = Decimal(0)
    for tw in tax_withheld_list:
        for t in tw.get("TaxesWithheld", []):
            t_amt = t.get("ChargeAmount") or {}
            amt = t_amt.get("CurrencyAmount", 0)
            tax_withheld += Decimal(str(amt))

    # Taxを除外して、商品価格と配送料のみを売上とする
    gross = charges["Principal"] + charges["ShippingCharge"]
    amazon_fee = sum(fees.values())  # マイナス値
    net = gross + amazon_fee

    return {
        "order_item_id": order_item_id,
        "principal": charges["Principal"],
        "shipping": charges["ShippingCharge"],
        "gross": gross,
        "amazon_fee": amazon_fee,
        "net": net,
    }

def build_finance_map(rows):
    fm = defaultdict(lambda: defaultdict(Decimal))
    for r in rows:
        oid = r["order_item_id"]
        if not oid:
            continue
        for k, v in r.items():
            if k != "order_item_id":
                fm[oid][k] += v
    return fm

# =========================
# Excel 出力
# =========================
def load_existing_data(path):
    """既存のExcelデータを読み込み、Order Item IDをキーにした辞書と、最新の受注日を返す"""
    data_map = {}
    manual_headers = []
    max_date = None
    
    if not path.exists():
        return data_map, manual_headers, max_date

    wb = load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        return data_map, manual_headers, max_date
    
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        return data_map, manual_headers, max_date

    # ヘッダー解析
    existing_headers = list(rows[0])
    
    # プログラム管理外のヘッダー（手入力列）を特定
    # buyer-name は削除対象なので除外
    manual_headers = [h for h in existing_headers if h not in HEADERS and h != "buyer-name"]
    
    # 列インデックスのマップ作成
    header_idx = {h: i for i, h in enumerate(existing_headers)}
    
    for r in rows[1:]:
        row_dict = {h: r[i] for h, i in header_idx.items() if i < len(r)}
        
        oid = row_dict.get("order-item-id")
        if not oid:
            continue
            
        # 日付の最大値を取得
        p_date_str = row_dict.get("purchase-date")
        if p_date_str:
            try:
                # 文字列またはdatetime型に対応
                if isinstance(p_date_str, str):
                    dt = parser.parse(p_date_str)
                else:
                    dt = p_date_str
                
                # タイムゾーンがない場合はUTCとみなす（比較用）
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                
                if max_date is None or dt > max_date:
                    max_date = dt
            except:
                pass

        data_map[oid] = row_dict

    return data_map, manual_headers, max_date

def update_excel(new_rows, fee_rows, existing_data, manual_headers):
    path = Path(EXCEL_PATH)
    
    # 既存データと新規データをマージ
    # Order Item ID をキーにして統合
    merged_map = existing_data.copy()
    
    for row in new_rows:
        oid = row["order-item-id"]
        
        # 既存データがあり、かつ「Finance連携済み」なら更新しない（手入力データ保護のため）
        if oid in existing_data:
            status = existing_data[oid].get("finance-status")
            if status == "Done":
                continue
        
        # マージ（既存の手入力データがあれば保持）
        current_manual_data = {k: existing_data[oid][k] for k in manual_headers if k in existing_data[oid]} if oid in existing_data else {}
        row.update(current_manual_data)
        merged_map[oid] = row

    # 書き込み準備
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
    
    # --- Sheet 1: Orders ---
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row) # 一旦クリアして書き直す
    else:
        ws = wb.create_sheet(SHEET_NAME)
        if "Sheet" in wb.sheetnames: del wb["Sheet"]

    # ヘッダー書き込み (プログラム管理列 + 手入力列)
    all_headers = HEADERS + manual_headers
    ws.append(all_headers)

    for c in range(1, len(all_headers)+1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    # データ書き込み（日付順にソート）
    # ソートキー: purchase-date (文字列またはdatetime)
    def get_sort_key(item):
        val = item.get("purchase-date")
        if not val: return ""
        return str(val)

    sorted_rows = sorted(merged_map.values(), key=get_sort_key, reverse=False)

    for r in sorted_rows:
        row_values = []
        for h in all_headers:
            val = r.get(h)
            # 数値型への変換が必要なカラム
            if h in ["principal", "shipping-price", "gross-sales", "amazon-fee", "net-payout"] and val is not None:
                try:
                    val = float(val)
                except:
                    pass
            row_values.append(val)
        ws.append(row_values)

    # --- Sheet 2: Service Fees ---
    FEE_SHEET = "ServiceFees"
    ws_fee = wb[FEE_SHEET] if FEE_SHEET in wb.sheetnames else wb.create_sheet(FEE_SHEET)
    ws_fee.delete_rows(1, ws_fee.max_row)

    headers_fee = ["date", "type", "description", "amount"]
    ws_fee.append(headers_fee)
    for c in range(1, len(headers_fee)+1):
        ws_fee.cell(row=1, column=c).font = Font(bold=True)

    for r in fee_rows:
        ws_fee.append([r["date"], r["type"], r["description"], float(r["amount"])])

    wb.save(path)

# =========================
# main
# =========================
if __name__ == "__main__":
    # 1. 既存データの読み込み
    print("[INFO] Loading existing Excel data...")
    existing_data, manual_headers, max_date = load_existing_data(Path(EXCEL_PATH))
    
    # 2. 取得開始日の決定 (最新日付 - 1日)
    if max_date:
        start_date = max_date - timedelta(days=1)
        created_after = start_date.strftime("%Y-%m-%dT%H:%M:%SZ")
        print(f"[INFO] Fetching orders after: {created_after} (Based on Excel data)")
    else:
        created_after = "2026-01-01T00:00:00Z"
        print(f"[INFO] Fetching orders after: {created_after} (Default)")

    orders = fetch_orders(created_after)
    finance_rows, service_fee_rows = fetch_finances()
    finance_map = build_finance_map(finance_rows)

    token = get_lwa_token()
    auth = get_auth()

    excel_rows = []
    current_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for o in orders:
        # 既存データで「Done」になっている注文はAPI詳細取得をスキップ（高速化）
        # ただし、ordersリストには入っているので、ここでチェック
        # ※ fetch_orders は期間指定で取ってくるため、Doneのものも含まれる可能性がある
        # ここでスキップしても update_excel で既存データが使われるので問題ない
        # (order-item-id単位の判定は後で行うが、APIコール節約のためOrder単位でチェックしたいが、Item単位のDone状況が混在する可能性は低い)
        
        # APIレートリミット対策 (1秒待機)
        time.sleep(1.0)
        try:
            items = fetch_order_items(o["AmazonOrderId"], token, auth)
        except Exception as e:
            print(f"[ERROR] fetch_order_items failed for {o['AmazonOrderId']}: {e}")
            continue

        addr = o.get("ShippingAddress")
        buyer_info = o.get("BuyerInfo") or {}

        for it in items:
            oid = it["OrderItemId"]
            f = finance_map.get(oid, {})

            # --- 金額取得ロジック (Finance API優先、なければOrders API) ---
            principal = f.get("principal", Decimal("0"))
            shipping = f.get("shipping", Decimal("0"))
            gross = f.get("gross", Decimal("0"))

            # Finance情報(出荷後)がない場合、Orders API(注文時)の情報を使用する
            if principal == 0:
                if "ItemPrice" in it and it["ItemPrice"]:
                    principal = Decimal(str(it["ItemPrice"].get("Amount", "0")))
                if "ShippingPrice" in it and it["ShippingPrice"]:
                    shipping = Decimal(str(it["ShippingPrice"].get("Amount", "0")))
                
                # grossを再計算
                if gross == 0:
                    gross = principal + shipping
            # -------------------------------------------------------

            amazon_fee = f.get("amazon_fee", Decimal("0"))
            
            # 手数料が0（未確定）の場合の概算取得
            if amazon_fee == 0 and principal > 0:
                is_fba = (it.get("FulfillmentChannel") == "AFN")
                curr = "USD"
                if "ItemPrice" in it and it["ItemPrice"]:
                    curr = it["ItemPrice"].get("CurrencyCode", "USD")
                elif o.get("OrderTotal"):
                    curr = o.get("OrderTotal", {}).get("CurrencyCode", "USD")
                
                est_fee = get_fees_estimate(it.get("ASIN"), principal, curr, is_fba, token, auth)
                if est_fee != 0:
                    amazon_fee = est_fee

            net = f.get("net", Decimal("0"))
            if net == 0 and (principal > 0 or amazon_fee != 0):
                net = gross + amazon_fee
            
            # Finance Status の判定
            # Finance APIからデータが取得できている場合のみ Done とする
            # (概算値の場合は Done にしないことで、次回以降も更新対象にする)
            finance_status = "Done" if oid in finance_map else ""

            excel_rows.append({
                "amazon-order-id": o["AmazonOrderId"],
                "purchase-date": o["PurchaseDate"],
                "order-status": o["OrderStatus"],
                "asin": it.get("ASIN"),
                "quantity": it.get("QuantityOrdered"),
                "principal": principal,
                "shipping": shipping,
                "gross": gross,
                "amazon_fee": amazon_fee,
                "net": net,
                "ship-city": addr.get("City") if addr else None,
                "ship-state": addr.get("StateOrRegion") if addr else None,
                "ship-postal-code": addr.get("PostalCode") if addr else None,
                "ship-country": addr.get("CountryCode") if addr else None,
                "recipient-name": addr.get("Name") if addr else None,
                "currency": o.get("OrderTotal", {}).get("CurrencyCode"),
                "order-item-id": oid,
                "fulfillment": it.get("FulfillmentChannel"),
                "finance-status": finance_status,
                "last-updated": current_time_str,
            })

    try:
        update_excel(excel_rows, service_fee_rows, existing_data, manual_headers)
        print(f"[OK] Excel updated: Total Orders={len(existing_data) + len(excel_rows)} (approx), Fees={len(service_fee_rows)} rows")
    except PermissionError:
        print(f"[ERROR] Excelファイルが開かれているため保存できません: {EXCEL_PATH}")
    except Exception as e:
        print(f"[ERROR] Excel write failed: {e}")
