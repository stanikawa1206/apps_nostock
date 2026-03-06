import time
import glob
import os
import re
import pandas as pd
import win32com.client as win32
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from apps.common.utils import (
    get_sql_server_connection,
)

def build_driver_for_ebay_csv():

    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options

    opts = Options()

    # ★ VBAと同じ
    opts.add_argument(
        r"--user-data-dir=C:\Users\stani\AppData\Local\Google\Chrome\User Data"
    )

    opts.add_argument("--profile-directory=Default")

    opts.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(options=opts)

    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )

    return driver

# ==========================================
# アカウント取得
# ==========================================
def get_accounts():
    conn = get_sql_server_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT account, username, [password]
        FROM mst.ebay_accounts
        WHERE ISNULL(is_excluded,0) = 0           
    """)

    rows = cur.fetchall()
    conn.close()
    return rows


# ==========================================
# ログイン（必要時のみ）
# ==========================================
from selenium.common.exceptions import NoSuchElementException
import time

def ensure_login(driver, username, password):

    driver.get("https://signin.ebay.com/signin/")
    time.sleep(3)

    print("ログイン画面確認中...")

    # 最大5分待機
    timeout = time.time() + 300

    while time.time() < timeout:

        # ① userid があれば通常ログイン
        try:
            user_input = driver.find_element(By.ID, "userid")
            user_input.clear()
            user_input.send_keys(username)

            driver.find_element(By.ID, "signin-continue-btn").click()
            time.sleep(2)

            pwd_input = driver.find_element(By.XPATH, "//input[@type='password']")
            pwd_input.clear()
            pwd_input.send_keys(password)

            driver.find_element(By.ID, "sgnBt").click()
            print("ログイン実行")
            time.sleep(5)
            return

        except NoSuchElementException:
            pass

        # ② CAPTCHA判定
        if "verify" in driver.page_source.lower():
            print("CAPTCHA発生 → 手動で対応してください")
            time.sleep(5)
            continue

        time.sleep(2)

    raise Exception("ログイン画面に到達できませんでした")


def get_account_snapshot(driver):

    driver.get("https://www.ebay.com/sh/ovw")
    time.sleep(5)

    result = {
        "free_left": None,
        "listed_count": None,
        "amount_used": None,
        "amount_left": None
    }

    try:
        # =========================
        # ① Premium Free Left
        # =========================
        promo_text = driver.find_element(
            By.XPATH,
            "//*[contains(text(),'Used/Left')]"
        ).text

        # 例: Used/Left: 6,646 / 3,354
        import re
        m = re.search(r'(\d{1,3}(?:,\d{3})*)\s*/\s*(\d{1,3}(?:,\d{3})*)', promo_text)

        if m:
            result["free_left"] = int(m.group(2).replace(",", ""))

    except:
        pass

    try:
        # =========================
        # ② 出品数
        # =========================
        qty_text = driver.find_element(
            By.XPATH,
            "//*[contains(text(),'listed and sold') and contains(text(),'limit on quantity')]"
        ).text

        # 例: 5,094 listed and sold / 100,000 limit on quantity of items
        m = re.search(r'(\d{1,3}(?:,\d{3})*)\s+listed', qty_text)
        if m:
            result["listed_count"] = int(m.group(1).replace(",", ""))

    except:
        pass

    try:
        # =========================
        # ③ 空き金額
        # =========================
        money_left_text = driver.find_element(
            By.XPATH,
            "//*[contains(text(),'more') and contains(text(),'$')]"
        ).text

        # 例: $2,430.15 more
        m = re.search(r'\$([\d,]+\.\d+)', money_left_text)
        if m:
            result["amount_left"] = float(m.group(1).replace(",", ""))

    except:
        pass

    try:
        # =========================
        # ④ 使用金額
        # =========================
        amount_text = driver.find_element(
            By.XPATH,
            "//*[contains(text(),'listed and sold') and contains(text(),'limit')]"
        ).text

        # 例: $2.4M listed and sold / $2.4M limit
        m = re.search(r'\$([\d\.]+)M\s+listed', amount_text)

        if m:
            result["amount_used"] = float(m.group(1)) * 1_000_000

    except:
        pass

    return result


# ==========================================
# Excel書き込み
# ==========================================
def write_excel(account_name,
                free_left,
                listed_count,
                amount_used,
                amount_left):

    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False

    wb = excel.Workbooks.Open(r"Y:\ebay\ebay.xlsx")
    ws = wb.Sheets("アカウント別")

    for col in range(1, 20):
        if ws.Cells(1, col).Value == account_name:

            ws.Cells(2, col).Value = free_left
            ws.Cells(3, col).Value = listed_count
            ws.Cells(4, col).Value = amount_used
            ws.Cells(5, col).Value = amount_left

            break

    wb.Save()
    wb.Close()
    excel.Quit()



# ==========================================
# レポートDL
# ==========================================
def download_report(driver, download_dir):

    wait = WebDriverWait(driver, 30)

    driver.get("https://www.ebay.com/sh/reports/downloads")

    # ───────── RefID取得待機 ─────────
    wait.until(
        EC.presence_of_element_located((By.XPATH, "//td[3]//div"))
    )

    ref_elements = driver.find_elements(By.XPATH, "//td[3]//div")
    initial_ref = None
    for el in ref_elements:
        txt = el.text.strip()
        if txt.isdigit() and len(txt) >= 8:
            initial_ref = txt
            break

    if not initial_ref:
        raise Exception("初期RefID取得失敗")

    # ───────── Download report ─────────
    download_btn = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(., 'Download report')]")
        )
    )
    download_btn.click()

    # ───────── LISTINGS ─────────
    select_source = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, "//div[text()='Select report source']/ancestor::button")
        )
    )

    driver.execute_script("arguments[0].click();", select_source)

    # ───────── ALL ACTIVE ─────────
    listings_radio = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, "//label[text()='Listings']")
        )
    )

    driver.execute_script("arguments[0].click();", listings_radio)

    report_type = wait.until(
        EC.presence_of_element_located(
            (By.XPATH, "//div[text()='Select report type']/ancestor::button")
        )
    )

    driver.execute_script("arguments[0].click();", report_type)

    all_active_radio = wait.until(
        EC.presence_of_element_located(
            (By.XPATH, "//label[text()='All active listings']")
        )
    )

    driver.execute_script("arguments[0].click();", all_active_radio)

    # ───────── 最終Download ─────────
    download_btn = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, "//div[contains(@class,'lightbox-dialog')]//button[text()='Download']")
        )
    )

    driver.execute_script("arguments[0].click();", download_btn)

    # ───────── 新RefID検出 ─────────
    wait = WebDriverWait(driver, 60)

    def get_new_ref(d):
        el = d.find_element(
            By.XPATH,
            "(//td[contains(@class,'requestId')]//div)[1]"
        )
        txt = el.text.strip()

        if txt.isdigit() and txt != initial_ref:
            return txt
        return False

    new_ref = wait.until(get_new_ref)

    print("新RefID:", new_ref)
    
    # ───────── CSV検出 ─────────
    timeout = time.time() + 180
    while time.time() < timeout:
        for f in os.listdir(download_dir):
            if new_ref in f and f.endswith(".csv"):
                return os.path.join(download_dir, f)
        time.sleep(1)

    raise Exception("CSV未検出")




# ==========================================
# 新規CSV特定
# ==========================================
def get_new_csv(before_files):

    path = r"C:\Users\stani\Downloads\*.csv"
    after_files = set(glob.glob(path))

    new_files = list(after_files - before_files)

    if not new_files:
        raise Exception("新しいCSVが見つかりません")

    return new_files[0]


# ==========================================
# SQL書込（高速版）
# ==========================================
def write_to_sql(account, df):

    # ←ここで型を整える
    df["Start price"] = pd.to_numeric(df["Start price"], errors="coerce").fillna(0)
    df["Watchers"] = pd.to_numeric(df["Watchers"], errors="coerce").fillna(0)

    conn = get_sql_server_connection()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM ext.ebay_active_download
        WHERE account = ?
    """, account)

    conn.commit()

    data = [
        (
            account,
            row["Item number"],
            row["Custom label (SKU)"],
            row["Title"],
            row.get("Watchers", 0),
            row["Start price"],
        )
        for _, row in df.iterrows()
    ]

    cur.executemany("""
        INSERT INTO ext.ebay_active_download
        (account, listing_id, vendor_item_id, title_en, watchers, [Start price])
        VALUES (?, ?, ?, ?, ?, ?)
    """, data)

    conn.commit()
    conn.close()

import time
import re
from openpyxl import load_workbook
from selenium.webdriver.common.by import By


EXCEL_PATH = r"Y:\ebay\ebay.xlsm"



def write_extract_to_excel(account, promo_used, selling_limit1, selling_limit2, selling_limit3):

    wb = load_workbook(EXCEL_PATH, keep_vba=True)
    ws = wb["アカウント別"]

    target_col = None

    # 1行目からアカウント列を探す
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == account:
            target_col = col
            break

    if target_col is None:
        raise Exception(f"{account} の列が見つかりません")

    # 2〜5行クリア
    for r in range(2, 6):
        ws.cell(r, target_col).value = None

    # VBAと同じ場所に書く
    ws.cell(2, target_col).value = promo_used
    ws.cell(3, target_col).value = selling_limit1
    ws.cell(4, target_col).value = selling_limit2
    ws.cell(5, target_col).value = selling_limit3

    wb.save(EXCEL_PATH)
    wb.close()

def ebay_extract_data(driver, account):

    promo_used = None
    selling_limit1 = ""
    selling_limit2 = ""
    selling_limit3 = ""

    # ① ページロード待ち
    timeout = time.time() + 10
    while time.time() < timeout:
        state = driver.execute_script("return document.readyState")
        if state == "complete":
            break
        time.sleep(0.2)

    # ② show more クリック
    try:
        promo_btn = driver.find_element(
            By.XPATH,
            "//*[@id='sho-prom-offers']/div/div[2]/div/div/div[8]/div/button"
        )
        promo_btn.click()
    except Exception:
        pass

    time.sleep(2)

    # ③ HTML取得
    page_content = driver.execute_script("return document.body.innerHTML")

    # ④ Promotional offers used 抽出
    pos = page_content.find("Premium Store Subscription")
    if pos != -1:
        promo_pos = page_content.find("Promotional offers,", pos)
        if promo_pos != -1:
            used_pos = page_content.find("used,", promo_pos)
            if used_pos != -1:
                number_text = page_content[used_pos + len("used,"): used_pos + len("used,") + 20]
                m = re.search(r"\d+", number_text)
                if m:
                    promo_used = m.group()

    # ⑤ Selling limits取得
    timeout = time.time() + 5
    while time.time() < timeout:
        try:
            selling_limit1 = driver.find_element(
                By.XPATH,
                "//*[@id='sho-selling-limits']/div/div[2]/div/div/div[2]/p/span[1]"
            ).text

            selling_limit2 = driver.find_element(
                By.XPATH,
                "//*[@id='sho-selling-limits']/div/div[2]/div/div/div[1]/p/span[1]"
            ).text

            selling_limit3_raw = driver.find_element(
                By.XPATH,
                "//*[@id='sho-selling-limits']/div/div[2]/div/div/div[2]/h3/span"
            ).text

            if selling_limit1 and selling_limit2 and selling_limit3_raw:
                m = re.search(r"\d{1,3}(?:,\d{3})*(?:\.\d+)?", selling_limit3_raw)
                if m:
                    selling_limit3 = m.group()
                break

        except Exception:
            pass

        time.sleep(0.5)

    # Excelへ書く前に確認
    print("account:", account)
    print("promo_used:", promo_used)
    print("selling_limit1:", selling_limit1)
    print("selling_limit2:", selling_limit2)
    print("selling_limit3:", selling_limit3)

    # Excelへ直接書き込み
    write_extract_to_excel(
        account=account,
        promo_used=promo_used,
        selling_limit1=selling_limit1,
        selling_limit2=selling_limit2,
        selling_limit3=selling_limit3
    )

# ==========================================
# メイン
# ==========================================
def main():

    accounts = get_accounts()

    for account, username, password in accounts:

        print(f"==== {account} 開始 ====")

        driver = build_driver_for_ebay_csv()

        ensure_login(driver, username, password)

        ebay_extract_data(driver,account)

        return


        # CSVだけ
        before_files = set(glob.glob(r"C:\Users\stani\Downloads\*.csv"))

        download_dir = r"C:\Users\stani\Downloads"
        download_report(driver, download_dir)

        csv_file = get_new_csv(before_files)

        df = pd.read_csv(csv_file)

        write_to_sql(account, df)

        driver.quit()

        print(f"==== {account} 完了 ====")


if __name__ == "__main__":
    main()
